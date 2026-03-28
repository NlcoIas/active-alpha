"""State Street Global Advisors (SPDR) ETF holdings client.

Fetches daily holdings XLSX files published by SSGA for their SPDR ETFs.
"""

from __future__ import annotations

import io
import logging
from typing import Any

import httpx

logger = logging.getLogger(__name__)

SSGA_URL_TEMPLATE = (
    "https://www.ssga.com/library-content/products/fund-data/etfs/us/"
    "holdings-daily-us-en-{ticker}.xlsx"
)

# Known SPDR ETF tickers — can be expanded as needed
SSGA_TICKERS: set[str] = {
    # S&P 500 family
    "SPY", "SPYG", "SPYV", "SPYD",
    # Select Sector SPDRs
    "XLK", "XLF", "XLE", "XLV", "XLI", "XLU", "XLC", "XLRE", "XLB", "XLP", "XLY",
    # International
    "SPDW", "SPEM",
    # Fixed income
    "BIL", "SPAB", "SPTS", "SPTL", "SPIB",
    # Other popular
    "GLD", "MDYG", "MDYV", "SLYG", "SLYV",
    "SDY", "SPMD", "SPSM",
    "KBE", "KRE", "XBI", "XHB", "XME", "XOP", "XRT",
}

# Column name aliases for flexible header matching
_COLUMN_ALIASES: dict[str, list[str]] = {
    "ticker": ["ticker", "symbol", "stock ticker"],
    "name": ["name", "security name", "description", "holding name"],
    "weight_pct": ["weight", "weight (%)", "% of net assets", "portfolio weight"],
    "shares": ["shares held", "shares", "quantity", "shares/par amount"],
    "market_value": ["market value", "emv", "market value ($)", "value"],
}


def _find_column_index(
    header_row: list[Any], aliases: list[str]
) -> int | None:
    """Find column index by matching header cell text against aliases."""
    for idx, cell in enumerate(header_row):
        cell_text = str(cell).strip().lower() if cell is not None else ""
        for alias in aliases:
            if alias.lower() == cell_text:
                return idx
    return None


def _safe_float(value: Any) -> float | None:
    """Parse a value to float, handling strings with commas/symbols."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).strip().replace(",", "").replace("$", "").replace("%", "")
    if not cleaned or cleaned == "-" or cleaned.lower() in ("n/a", "na", "--"):
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


class SSGAClient:
    """Client for fetching SPDR ETF holdings from State Street's public XLSX files."""

    def __init__(self, timeout: float = 30.0) -> None:
        self._http = httpx.AsyncClient(
            timeout=timeout,
            headers={"User-Agent": "ActiveAlpha/1.0 (ETF research tool)"},
        )

    def supports(self, ticker: str) -> bool:
        """Check if this provider can fetch holdings for this ticker."""
        return ticker.upper() in SSGA_TICKERS

    async def fetch_holdings(self, ticker: str) -> list[dict[str, Any]]:
        """Fetch holdings for an SPDR ETF.

        Returns list of dicts with keys: ticker, name, weight_pct, shares, market_value.
        Weight_pct is in percentage points (e.g. 7.24 means 7.24%).
        Returns empty list on any error.
        """
        ticker_lower = ticker.lower()
        url = SSGA_URL_TEMPLATE.format(ticker=ticker_lower)

        try:
            response = await self._http.get(url)
            response.raise_for_status()
        except httpx.HTTPStatusError as exc:
            logger.error(
                "SSGA: HTTP %d fetching %s holdings: %s",
                exc.response.status_code,
                ticker.upper(),
                exc,
            )
            return []
        except httpx.HTTPError as exc:
            logger.error("SSGA: network error fetching %s holdings: %s", ticker.upper(), exc)
            return []

        try:
            return self._parse_xlsx(response.content, ticker.upper())
        except Exception:
            logger.exception("SSGA: failed to parse XLSX for %s", ticker.upper())
            return []

    def _parse_xlsx(self, content: bytes, ticker: str) -> list[dict[str, Any]]:
        """Parse SSGA XLSX bytes into normalized holdings list."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error("SSGA: openpyxl is required but not installed")
            return []

        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            logger.warning("SSGA: no active worksheet in XLSX for %s", ticker)
            wb.close()
            return []

        # Find the header row by scanning for a row containing "Ticker" or "Name"
        header_row_idx: int | None = None
        header_cells: list[Any] = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            row_values = list(row)
            text_values = [
                str(cell).strip().lower() for cell in row_values if cell is not None
            ]
            if any(
                val in text_values
                for val in ("ticker", "name", "symbol", "security name")
            ):
                header_row_idx = row_idx
                header_cells = row_values
                break

        if header_row_idx is None:
            logger.warning(
                "SSGA: could not find header row in XLSX for %s", ticker
            )
            wb.close()
            return []

        # Map columns
        col_ticker = _find_column_index(header_cells, _COLUMN_ALIASES["ticker"])
        col_name = _find_column_index(header_cells, _COLUMN_ALIASES["name"])
        col_weight = _find_column_index(header_cells, _COLUMN_ALIASES["weight_pct"])
        col_shares = _find_column_index(header_cells, _COLUMN_ALIASES["shares"])
        col_market_value = _find_column_index(header_cells, _COLUMN_ALIASES["market_value"])

        if col_ticker is None and col_name is None:
            logger.warning(
                "SSGA: could not identify ticker or name columns for %s. Headers: %s",
                ticker,
                header_cells,
            )
            wb.close()
            return []

        # Read data rows after the header
        holdings: list[dict[str, Any]] = []
        data_started = False

        for row in ws.iter_rows(values_only=True):
            row_values = list(row)
            if not data_started:
                # Skip rows until we pass the header
                text_values = [
                    str(cell).strip().lower() for cell in row_values if cell is not None
                ]
                if any(
                    val in text_values
                    for val in ("ticker", "name", "symbol", "security name")
                ):
                    data_started = True
                continue

            # Skip fully empty rows
            if all(cell is None or str(cell).strip() == "" for cell in row_values):
                continue

            holding_ticker = ""
            if col_ticker is not None and col_ticker < len(row_values):
                val = row_values[col_ticker]
                holding_ticker = str(val).strip() if val is not None else ""

            holding_name = ""
            if col_name is not None and col_name < len(row_values):
                val = row_values[col_name]
                holding_name = str(val).strip() if val is not None else ""

            # Skip rows with no meaningful data
            if not holding_ticker and not holding_name:
                continue

            weight_raw = None
            if col_weight is not None and col_weight < len(row_values):
                weight_raw = _safe_float(row_values[col_weight])

            # SSGA weights are decimals (0.0724 = 7.24%) — convert to percentage
            weight_pct: float | None = None
            if weight_raw is not None:
                if abs(weight_raw) <= 1.0:
                    weight_pct = weight_raw * 100.0
                else:
                    # Already in percentage form
                    weight_pct = weight_raw

            shares: float | None = None
            if col_shares is not None and col_shares < len(row_values):
                shares = _safe_float(row_values[col_shares])

            market_value: float | None = None
            if col_market_value is not None and col_market_value < len(row_values):
                market_value = _safe_float(row_values[col_market_value])

            holdings.append(
                {
                    "ticker": holding_ticker,
                    "name": holding_name,
                    "weight_pct": weight_pct,
                    "shares": shares,
                    "market_value": market_value,
                }
            )

        wb.close()
        logger.info("SSGA: parsed %d holdings for %s", len(holdings), ticker)
        return holdings

    async def close(self) -> None:
        """Close the underlying HTTP client."""
        await self._http.aclose()

    async def __aenter__(self) -> SSGAClient:
        return self

    async def __aexit__(self, *args: Any) -> None:
        await self.close()
